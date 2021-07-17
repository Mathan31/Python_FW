import openpyxl

class Excel:

    def get_total_row(self,path,sheetname):
        wb = openpyxl.open(path)
        sheet = wb[sheetname]
        return sheet.max_row

    def get_total_col(self,path,sheetname):
        wb = openpyxl.open(path)
        sheet = wb[sheetname]
        return sheet.max_column

    def get_value(self,path,sheetname,row,col):
        wb = openpyxl.open(path)
        sheet = wb[sheetname]
        return sheet.cell(row,col).value

    def set_value(self,path,sheetname,row,col,data):
        wb = openpyxl.open(path)
        sheet = wb[sheetname]
        sheet.cell(row, col).value = data
        wb.save(path)

        